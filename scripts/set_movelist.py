"""
This is used to update the movelists of all the monsters.
Put the techniques.xlsx file in the scripts folder alongside this.
The movelists are expected to be in "Sheet4". There should be a row of mon names, each having three moves
underneath.
Call this script from the root folder.
Any missing movelists, or missing json files will be logged to stdout and to a file.
"""

import glob
import re
from json import loads
from openpyxl import load_workbook


PRINT_SUMMARY = True
excel_path = 'scripts/techniques.xlsx'
wb2 = load_workbook(excel_path)
move_list = wb2.get_sheet_by_name('Sheet4')
defaults = {
    "Water": [
        "Flood",
        "Flow",
        "Font",
    ],
    "Wood": [
        "Fluff Up",
        "Sting",
        "Splinter",
    ],
    "Earth": [
        "Mudslide",
        "Ram",
        "Rock",
    ],
    "Fire": [
        "Fire Ball",
        "Fire Claw",
        "Flamethrower",
    ],
    "Metal": [
        "Perfect Cut",
        "Shrapnel",
        "Wall of Steel",
    ]
}

missing_json_file = set()
mons = {}
name_trans = {}
updated_mons = set()
mons_without_a_moveset = set()


def set_moves(f, data, moves):
    moves_string = '''"moveset": [
        {
            "level_learned": 2,
            "technique": "%s"
        },
        {
            "level_learned": 2,
            "technique": "%s"
        },
        {
            "level_learned": 2,
            "technique": "%s"
        }
    ],''' % (
        moves[0],
        moves[1],
        moves[2],
    )
    out = re.sub(r'"moveset": \[(.*?)\],', moves_string, data, flags=re.S)
    f.truncate(0)
    f.write(out)


def slugify(data):
    return "%s" % (data.lower().replace(" ", "_").replace("-", "_"))


for x in range(0, 2000):
    try:
        row = [move_list.cell(row=y, column=x).value for y in range(2, 6)]
    except IndexError:
        break
    if None in row:
        continue
    mons[row[0]] = [
        slugify(attack)
        for attack in row[1:]
    ]


for mon, moves in mons.items():
    path = "tuxemon/resources/db/monster/%s.json" % mon.lower()
    data = None
    try:
        with open(path, "r") as f:
            data = f.read()
            updated_mons.add(path)
    except IOError:
        missing_json_file.add(mon)
        continue

    with open(path, "w") as f:
        set_moves(f, data, moves)


if missing_json_file:
    path = "mons_without_a_json_file.txt"
    print("%s mons have movelists, but don't have a json file. See %s for a summary" % (
        len(missing_json_file),
        path,
    ))
    print(missing_json_file)
    if PRINT_SUMMARY:
        with open(path, "w") as f:
            for mon in missing_json_file:
                f.write(mon + "\n")

monster_path = "tuxemon/resources/db/monster/*.json"
jsons = glob.glob(monster_path)
for path in jsons:
    if path not in updated_mons:
        with open(path, "r") as f2:
            data = f2.read()
            t = loads(data)['types'][0]
            moves = [slugify(m) for m in defaults[t]]

        with open(path, "w") as f2:
            set_moves(f2, data, moves)

        mon = path.rpartition("/")[2].partition(".")[0]
        mons_without_a_moveset.add(mon)

if mons_without_a_moveset:
    path = "mons_without_a_moveset.txt"
    print(
            "%s existing mons didn't have a movelist defined in the excel doc. "
            "Used default moveset for type. See %s for a summary" % (
                len(mons_without_a_moveset),
                path,
            )
    )
    print(mons_without_a_moveset)
    if PRINT_SUMMARY:
        with open(path, "w") as f:
            for mon in missing_json_file:
                f.write(mon + "\n")
