"""
This is used to create the json files for the techniques.
Put the techniques.xlsx file in the scripts folder alongside this.
Call this script from the root folder.
"name_trans.txt" will also be created, which contains the english locale entries, and can
be copy-pasted into en_US.json
"""

from collections import namedtuple
from json import dump
from openpyxl import load_workbook

excel_path = 'scripts/techniques.xlsx'
wb2 = load_workbook(excel_path)
tech_sheet = wb2.get_sheet_by_name('Techs')

DataRow = namedtuple("DataRow", [
    "name",
    "id",
    "element",
    "recharge",
    "accuracy",
    "potency",
    "user_condition",
    "target_condition",
    "animation",
    "animation_target",
    "range",
    "power",
    "healing_power",
    "is_fast",
])

axes_flip_mapping = {
    (
        "hits_for_separation", "breath_blue", "breath_fire", "claw_blue",
        "claw_yellow_169", "fireball_114", "firelion_right", "lightning_bolt_138",
        "metal_delete", "power_arc_154", "pushtrap_right", "shield_turtle_right",
        "slash_200", "slash_fire", "screen", "snake_right", "tornado_basic", 
        "tornado_volume", "watershot",
    ): "x",
    ("lance_ice", "triforce_163"): "xy"
}


def get_animation_flip_axes(animation_name: str) -> str:
    """Defines in which axes should the animation be flipped."""
    for names, axes in axes_flip_mapping.items():
        if animation_name in names:
            return axes
    return ""


def create_json(data_row):
    name = data_row.name.lower().replace(" ", "_").replace("-", "_")
    name_trans['technique_%s_name' % name] = data_row.name.strip()
    types = [t.strip().lower() for t in data_row.element.split(",")]
    effects = []
    if data_row.power:
        effects += ["damage"]

    template = {
        "slug": name,
        "use_tech": "combat_used_x",
        "use_success": None,
        "use_failure": "combat_miss",
        "animation": data_row.animation,
        "sfx": "blaster1.ogg",
        "icon": "",
        "accuracy": data_row.accuracy,
        "effects": effects,
        "flip_axes": get_animation_flip_axes(data_row.animation),
        "healing_power": float(data_row.healing_power) if data_row.healing_power else 0,
        "is_fast": bool(data_row.is_fast),
        "potency": data_row.potency,
        "power": float(data_row.power) if data_row.power else 0,
        "range": data_row.range.lower(),
        "recharge": int(data_row.recharge),
        "sort": "damage",
        "target": {
            "enemy_monster": True,
            "enemy_team": False,
            "enemy_trainer": False,
            "own_monster": True,
            "own_team": False,
            "own_trainer": False,
        },
        "types": types,
    }

    path = "tuxemon/resources/db/technique/%s.json" % name
    with open(path, "w") as f:
        dump(template, f, indent=2, separators=(",", ": "), sort_keys=True)
        f.write("\n")


name_trans = {}

for y in range(2, 6000):
    row = DataRow(*(tech_sheet.cell(row=y, column=x).value for x in range(1, 16)))
    if row.name is None:
        break
    create_json(row)

path = "name_trans.txt"
with open(path, "w") as f:
    dump(name_trans, f, indent=2, separators=(",", ": "), sort_keys=True)
